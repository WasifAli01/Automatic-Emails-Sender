# Automatic-Emails-Sender
'''Hey this is a simple GUI application for sending Emails automatically through your Gmail account. Kindly read read-me for more description.'''
import openpyxl as xl
import smtplib as smtp
import tkinter as tk
from tkinter import messagebox
from tkinter import ttk
from tkinter import filedialog


def main():

        # Sending Emails

    # Tkinter GUI design

    root = tk.Tk()
    root.title('Automate Emails')

    root.geometry("500x400")
    root.resizable(0, 0)

    # Top Design
    topLabel = tk.Label(text="Login To Your Account",
                        bg="#99ddf0", fg="#000000", font=("Georgia", 10))
    topLabel.grid(row=1, column=2, pady=5, ipadx=25, ipady=15)
    # Email, Password and Other Entry Widgets
    emailLabel = tk.Label(root, text="Email: ")
    emailLabel.grid(row=2, column=1)
    passwordLabel = tk.Label(root, text="Password: ")
    passwordLabel.grid(row=3, column=1)

    emailInput = tk.Entry(root, width=30, border=0, bg="#c8cbcc", fg="#000000")
    emailInput.grid(row=2, column=2, pady=5)
    passwordInput = tk.Entry(root, width=30, border=0,
                             bg="#c8cbcc", fg="#000000")
    passwordInput.grid(row=3, column=2, pady=5)

    MessageLabelTop = tk.Label(
        root, text="Enter Message Details", bg="#99ddf0", fg="#000000", font=("Georgia", 10))
    MessageLabelTop.grid(row=4, column=2, pady=3, padx=3, ipadx=20, ipady=10)

    subjectLabel = tk.Label(root, text='Enter A Subject For Message')
    subjectLabel.grid(row=5, column=1)
    subjectInput = tk.Entry(root, bg="#c8cbcc", fg="#000", width=30)
    subjectInput.grid(row=5, column=2)

    sheetNameLabel = tk.Label(root, text="Enter Sheet Name")
    sheetNameLabel.grid(row=7, column=1)

    sheetName = tk.Entry(root, border=0, bg="#c8cbcc", fg="#000", width=30)
    sheetName.grid(row=7, column=2)

    column_NumberLabel = tk.Label(root, text="Enter Column Number For Emails")
    column_Number = tk.Entry(root, width=30, border=0, bg="#c8cbcc", fg="#000")
    column_NumberLabel.grid(column=1, row=8)
    column_Number.grid(column=2, row=8)

    def sendEmails():
        fileName = filedialog.askopenfilename(
            initialdir='/', title="Choose A File", filetype=((".xlsx", "*.xlsx"), ("All Files", "*.*")))
        fileNameLabel = tk.Label(root, text="")
        fileNameLabel.grid(row=6, column=2)
        theFileName = fileNameLabel.configure(text=fileName)
    # Excel File Working
        try:
            gmail = smtp.SMTP_SSL("smtp.gmail.com", 465)
            # Logging In
            gmail.ehlo()
            gmail.login(emailInput.get(), passwordInput.get())
            messagebox.showinfo('Successful', 'Login Successful')
        except smtp.SMTPAuthenticationError:
            messagebox.showerror(
                'Failed', message="Unable To Login To Your Account.Kindly Check Your Email and Password and Try Again.\nElse allow less secure apps access in your account settings.")
        except smtp.socket.gaierror:
            messagebox.showerror(
                'Failed', 'Kindly Check Your Internet Connection And Try Again')
        try:
            workBook = xl.load_workbook(fileNameLabel.cget("text"))
            sheet = workBook[sheetName.get()]
        except KeyError:
            messagebox(f"{sheet} doesn't exists")

        currentRow = 1
        for email in range(1, sheet.max_row + 1):
            message = open("message.txt", "r")
            currentEmail = sheet.cell(
                row=currentRow, column=int(column_Number.get())).value
            gmail.sendmail(emailInput.get(
            ), currentEmail, f"Subject:{subjectInput.get()}\n {message.read()} ")
            print(f"Sending Email to {currentEmail}")
            currentRow += 1
        gmail.quit()
        messagebox.showinfo(
            'Success', 'Message sent to all Emails Successfully.')
     # Send email to mails in selected file
    sendEmailBtn = tk.Button(root, text="Login & Send Emails",
                             border=0, bg="#fff", fg="#000", command=sendEmails, highlightthickness=0)
    sendEmailBtn.config(bd="2")
    sendEmailBtn.grid(row=9, column=2, ipadx=10, ipady=10, pady=3)

    # Main Loop

    root.mainloop()


main()
